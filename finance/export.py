__author__ = 'shu'
from django.http import HttpResponse
from conf import settings

def export_csv(modeladmin, request, queryset):
    import csv
    from django.utils.encoding import smart_str
    response = HttpResponse(mimetype='text/csv')
    response['Content-Disposition'] = 'attachment; filename=mymodel.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"ID"),
        smart_str(u"Title"),
        smart_str(u"Description"),
    ])
    for obj in queryset:
        writer.writerow([
            smart_str(obj.pk),
            smart_str(obj.title),
            smart_str(obj.description),
        ])
    return response
export_csv.short_description = u"Export CSV"

def export_xls(modeladmin, request, queryset):
    import xlwt
    response = HttpResponse(mimetype='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s_vp.xls' % settings.COUNTRY
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet(settings.COUNTRY + "_vp")

    row_num = 0

    columns = [
        (u"month", 276),
        (u"Volatility Premium", 276),
        (u"Country", 276),
    ]

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    for col_num in xrange(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        # set column width
        ws.col(col_num).width = columns[col_num][1]

    font_style = xlwt.XFStyle()
    font_style.alignment.wrap = 1

    for obj in queryset:
        row_num += 1
        row = [
            obj.month,
            obj.vp,
            obj.country,
        ]
        for col_num in xrange(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response

export_xls.short_description = u"Export XLS"

def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.cell import get_column_letter
    response = HttpResponse(mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = settings.COUNTRY + "_vp"

    row_num = 0

    columns = [
        (u"ID", 15),
        (u"Title", 70),
        (u"Description", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.title,
            obj.description,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response

export_xlsx.short_description = u"Export XLSX"


def monthly_groups_export_xls(modeladmin, request, queryset):
    import xlwt
    response = HttpResponse(mimetype='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s_monthly_groups.xls' % settings.COUNTRY
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet(settings.COUNTRY + "_monthly_groups")

    row_num = 0

    columns = [
        (u"month", 276),
        (u"Type", 276),
        (u"Group Number", 276),
        (u"Average Returns", 276),
        (u"Country", 276),
    ]

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    for col_num in xrange(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        # set column width
        ws.col(col_num).width = columns[col_num][1]

    font_style = xlwt.XFStyle()
    font_style.alignment.wrap = 1

    for obj in queryset:
        row_num += 1
        row = [
            obj.month,
            obj.type,
            obj.group_number,
            obj.average_returns,
            obj.country,
        ]
        for col_num in xrange(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response

export_xls.short_description = u"Export XLS"


def monthly_groups_results_export_xls(modeladmin, request, queryset):
    import xlwt
    response = HttpResponse(mimetype='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s_monthly_groups_results.xls' % settings.COUNTRY
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet(settings.COUNTRY + "_monthly_groups_results")

    row_num = 0

    columns = [
        (u"Type", 276),
        (u"Group Number", 276),
        (u"Results", 276),
        (u"Country", 276),
    ]

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    for col_num in xrange(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        # set column width
        ws.col(col_num).width = columns[col_num][1]

    font_style = xlwt.XFStyle()
    font_style.alignment.wrap = 1

    for obj in queryset:
        row_num += 1
        row = [
            obj.type,
            obj.group_number,
            obj.result,
            obj.country,
        ]
        for col_num in xrange(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response

export_xls.short_description = u"Export XLS"